from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from .models import StudentProfile
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import ContestForm, DSAQuestionForm,LeetCodeProfileForm,SocialLinksForm
from .models import Question, Contest, ContestQuestion
import requests,math
from firebase_config import db


def is_admin(user):
    return user.is_superuser

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json,random
from .models import ContestParticipation, ParticipationQuestion, ContestQuestion, StudentProfile

def get_leetcode_score_from_url(profile_url):
    try:
        username = profile_url.strip('/').split('/')[-1]
        graphql_url = "https://leetcode.com/graphql/"
        headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": profile_url
        }
        query = """
        query {
          matchedUser(username: "%s") {
            submitStats {
              acSubmissionNum {
                difficulty
                count
              }
            }
          }
        }
        """ % username

        payload = {"query": query}
        response = requests.post(graphql_url, json=payload, headers=headers)

        if response.status_code != 200:
            return 0

        data = response.json()
        user = data.get("data", {}).get("matchedUser")

        if not user:
            return 0

        easy = medium = hard = 0
        for item in user['submitStats']['acSubmissionNum']:
            difficulty = item['difficulty']
            count = item['count']
            if difficulty == 'Easy':
                easy = count
            elif difficulty == 'Medium':
                medium = count
            elif difficulty == 'Hard':
                hard = count

        score = (1 * easy) + (3 * medium) + (5 * hard)
        return score

    except Exception as e:
        print(e)
        return 0

@login_required(login_url='login')
@csrf_exempt
def update_leetcode_score(request):
    profile = get_object_or_404(StudentProfile, user=request.user)

    if profile.leetcode_profile_url:
        username = profile.leetcode_profile_url.rstrip('/').split('/')[-1]

        graphql_url = "https://leetcode.com/graphql/"
        headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": f"https://leetcode.com/u/{username}/"
        }

        query = """
        query {
          matchedUser(username: "%s") {
            submitStats {
              acSubmissionNum {
                difficulty
                count
              }
            }
          }
        }
        """ % username

        payload = {"query": query}
        response = requests.post(graphql_url, json=payload, headers=headers)

        if response.status_code != 200:
            return JsonResponse({'success': False, 'message': 'Failed to fetch data.'})

        data = response.json()
        stats = data.get("data", {}).get("matchedUser", {}).get("submitStats", {}).get("acSubmissionNum", [])

        easy = medium = hard = 0
        for item in stats:
            if item['difficulty'] == 'Easy':
                easy = item['count']
            elif item['difficulty'] == 'Medium':
                medium = item['count']
            elif item['difficulty'] == 'Hard':
                hard = item['count']

        score = (1 * easy) + (3 * medium) + (5 * hard)

        # 🔥 Save all values to profile now
        profile.leetcode_score = score
        profile.leetcode_easy = easy
        profile.leetcode_medium = medium
        profile.leetcode_hard = hard
        profile.save()

        return JsonResponse({
            'success': True,
            'score': score,
            'easy': easy,
            'medium': medium,
            'hard': hard
        })

    return JsonResponse({'success': False, 'message': 'No profile URL provided'})

from django.shortcuts import render, redirect
from .forms import SocialLinksForm
from .models import StudentProfile

def update_social_links(request):
    profile = StudentProfile.objects.get(user=request.user)

    if request.method == 'POST':
        form = SocialLinksForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('dashboard')  # Adjust to your actual dashboard route
    else:
        form = SocialLinksForm(instance=profile)

    return render(request, 'update_social_links.html', {'form': form})



def student_profile(request, username):
    # Fetch target user and profile by username
    user = get_object_or_404(User, username=username)
    profile = get_object_or_404(StudentProfile, user=user)

    # Get all contest participations for this student
    # participations = ContestParticipation.objects.filter(student=profile)

    if request.method == 'POST':
        form = LeetCodeProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "LeetCode profile updated.")
            return redirect('student_profile',username=request.user.username)
    else:
        form = LeetCodeProfileForm(instance=profile)
    
    overall_easy = overall_medium = overall_hard = 0
    total = ours = 0
    all_contests = Contest.objects.all()
    participations = []
    for contest in all_contests:
        p, created = ContestParticipation.objects.get_or_create(student=profile, contest=contest)
        
        easy = medium = hard = 0
        for pq in p.questions.filter(is_selected=True):
            level = pq.contest_question.question.level
            if level == 'Easy':
                easy += 1
            elif level == 'Medium':
                medium += 1
            elif level == 'Hard':
                hard += 1

        p.easy_count = easy
        p.medium_count = medium
        p.hard_count = hard
        overall_easy += easy
        overall_medium += medium
        overall_hard += hard
        # Calculate max score for the contest
        max_score = sum(cq.score for cq in ContestQuestion.objects.filter(contest=p.contest))
        p.max_score = max_score
        total+= max_score
        ours += p.total_score
        # Compute accuracy %
        p.accuracy = (p.total_score / max_score * 100) if max_score else 0
        participations.append(p)

    progress_percent = round((ours / total * 100), 2) if total else 0
    topic_titles = []
    topic_progress = []

    for p in participations:
        topic_titles.append(p.contest.title)
        percent = round((p.total_score / p.max_score * 100), 2) if p.max_score else 0
        topic_progress.append(percent)
    
    def generate_colors(n):
        base_colors = ['#4fc3f7', '#ffca28', '#ef5350', '#66bb6a', '#9575cd', '#f06292',
                    '#7986cb', '#a1887f', '#ba68c8', '#ff8a65']
        if n <= len(base_colors):
            return base_colors[:n]
        else:
            return base_colors + [f'#{random.randint(0x100000, 0xFFFFFF):06x}' for _ in range(n - len(base_colors))]    
    topic_colors = generate_colors(len(topic_titles))
    zipped_topics = list(zip(topic_titles, topic_colors[:len(topic_titles)]))
    chart_width = max(len(topic_titles) * 100, 800)
    chart_height = max(len(topic_titles) * 40, 600)
    return render(request, 'student_profile.html', {
        'student': user,  # user object for full name / username etc.
        'profile': profile,  # StudentProfile for college/project code
        'participations': participations,
        'form': form,
        'leetcode_easy': profile.leetcode_easy,
        'leetcode_medium': profile.leetcode_medium,
        'leetcode_hard': profile.leetcode_hard,
        'overall_easy': overall_easy,
        'overall_medium': overall_medium,
        'overall_hard': overall_hard,
        'total_score' : total,
        'ours_score' : math.ceil(ours),
        'progress_percent': progress_percent,
        'topic_titles': topic_titles,
        'topic_progress': topic_progress,  
        'topic_colors': topic_colors,
        'zipped_topics': zipped_topics,
        'chart_height': chart_height,
        'chart_width': chart_width

    })


@login_required(login_url='login')
def update_participation(request):
    print("update_participation called by:", request.user.username)

    if request.method == 'POST':
        data = json.loads(request.body)
        user = request.user

        try:
            student = StudentProfile.objects.get(user=user)
        except StudentProfile.DoesNotExist:
            return JsonResponse({'error': 'No student profile found for user'}, status=400)

        contest_id = data.get('contest_id')
        question_id = data.get('question_id')
        passed = data.get('passed')
        total = data.get('total')
        is_selected = data.get('is_selected')

        participation, _ = ContestParticipation.objects.get_or_create(
            student=student,
            contest_id=contest_id
        )

        contest_question = ContestQuestion.objects.get(id=question_id)
        pq, _ = ParticipationQuestion.objects.get_or_create(
            participation=participation,
            contest_question=contest_question
        )

        pq.passed_testcases = passed
        pq.total_testcases = total
        pq.is_selected = is_selected
        pq.save()

        total_score = 0
        for q in participation.questions.all():
            if q.is_selected and q.total_testcases > 0:
                score = q.contest_question.score
                accuracy = q.passed_testcases / q.total_testcases
                total_score += accuracy * score

        participation.total_score = total_score
        participation.save()

        return JsonResponse({'success': True, 'total_score': total_score})

    return JsonResponse({'error': 'Invalid request'}, status=400)

# Remove a question from a contest
def remove_question_from_contest(request, contest_id, cq_id):
    cq = get_object_or_404(ContestQuestion, id=cq_id, contest_id=contest_id)
    cq.delete()
    return redirect('admin_contest_detail', contest_id=contest_id)

def admin_contest_detail(request, contest_id):
    contest = get_object_or_404(Contest, id=contest_id)
    contest_questions = ContestQuestion.objects.filter(contest=contest)
    return render(request, 'admin_contest_detail.html', {
        'contest': contest,
        'contest_questions': contest_questions
    })


# Add questions to a contest — show all available questions not in this contest
from django.shortcuts import render, get_object_or_404, redirect
from .models import Question, Contest, ContestQuestion

def add_questions_to_contest(request, contest_id):
    contest = get_object_or_404(Contest, id=contest_id)

    # Get all questions already in the contest
    existing_question_ids = ContestQuestion.objects.filter(contest=contest).values_list('question_id', flat=True)

    # Get all questions NOT already in this contest
    available_questions = Question.objects.exclude(id__in=existing_question_ids)

    if request.method == 'POST':
        selected_question_ids = request.POST.getlist('question')
        for qid in selected_question_ids:
            score = request.POST.get(f'score_{qid}', 0)
            question = get_object_or_404(Question, id=qid)
            ContestQuestion.objects.create(contest=contest, question=question, score=score)
        return redirect('admin_contest_detail', contest_id=contest.id)

    return render(request, 'add_questions_to_contest.html', {
        'contest': contest,
        'questions': available_questions
    })



# Delete a contest and its associated ContestQuestion entries
def delete_contest(request, contest_id):
    contest = get_object_or_404(Contest, id=contest_id)
    contest.delete()
    return redirect('view_contests')

@user_passes_test(is_admin, login_url='admin_login')
def add_contest(request):
    if request.method == 'POST':
        contest_form = ContestForm(request.POST)
        if contest_form.is_valid():
            contest = contest_form.save()

            question_ids = request.POST.getlist('question')

            for qid in question_ids:
                question = Question.objects.get(id=qid)
                score = int(request.POST.get(f'score_{qid}', 0))
                ContestQuestion.objects.create(
                    contest=contest,
                    question=question,
                    score=score
                )
            return redirect('view_contests')
        else:
            print(contest_form.errors)
    else:
        contest_form = ContestForm()

    questions = Question.objects.all()
    return render(request, 'add_contest.html', {
        'contest_form': contest_form,
        'questions': questions
    })

@user_passes_test(is_admin, login_url='admin_login')
def view_contests(request):
    contests = Contest.objects.all()
    return render(request, 'view_contests.html', {'contests': contests})

import json

@login_required(login_url='login')
def student_view_contest_detail(request, contest_id):
    contest = get_object_or_404(Contest, id=contest_id)
    contest_questions = ContestQuestion.objects.filter(contest=contest)
    total_score = sum(cq.score for cq in contest_questions)

    student_score = 0
    participation_data = {}

    try:
        student = StudentProfile.objects.get(user=request.user)
        participation, _ = ContestParticipation.objects.get_or_create(student=student, contest=contest)

        # Ensure ParticipationQuestion exists for every contest question
        for cq in contest_questions:
            ParticipationQuestion.objects.get_or_create(
                participation=participation,
                contest_question=cq
            )

        # Build participation data with string keys
        for pq in participation.questions.all():
            participation_data[str(pq.contest_question.id)] = {
                'passed': pq.passed_testcases,
                'total': pq.total_testcases,
                'is_selected': pq.is_selected,
            }

        student_score = participation.total_score

    except StudentProfile.DoesNotExist:
        pass

    participation_data_json = json.dumps(participation_data)

    return render(request, 'contest_detail.html', {
        'contest': contest,
        'contest_questions': contest_questions,
        'total_score': total_score,
        'student_score': student_score,
        'participation_data': participation_data_json  # pass JSON string now
    })



def delete_question(request, id):
    question = get_object_or_404(Question, id=id)
    question.delete()
    return redirect('view_questions')

@user_passes_test(is_admin, login_url='admin_login')
def view_questions(request):
    questions = Question.objects.all()
    return render(request, 'view_questions.html', {'questions': questions})

@user_passes_test(is_admin, login_url='admin_login')
def add_question(request):
    if request.method == 'POST':
        form = DSAQuestionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('view_questions')
    else:
        form = DSAQuestionForm()
    return render(request, 'add_question.html', {'form': form})

def student_view_contests(request):
    if not request.user.is_authenticated:
        return redirect('login')

    profile = get_object_or_404(StudentProfile, user=request.user)
    contests = Contest.objects.all()
    enriched_contests = []

    for contest in contests:
        participation, _ = ContestParticipation.objects.get_or_create(student=profile, contest=contest)
        max_score = sum(q.score for q in ContestQuestion.objects.filter(contest=contest))
        student_score = participation.total_score
        accuracy = round((student_score / max_score * 100), 2) if max_score else 0

        # Add progress field to contest object (or use a wrapper dict)
        contest.progress = accuracy
        contest.remaining_progress = 100 - accuracy  # For chart logic
        enriched_contests.append(contest)

    return render(request, 'view_all_contests.html', {
        'contests': enriched_contests,
    })


def index(request):
    return render(request, 'index.html')

def register(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']
        full_name = request.POST['full_name']
        college_name = request.POST['college_name']
        project_code = request.POST['project_code']

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('register')

        if password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect('register')

        # Create user
        user = User.objects.create_user(username=username, password=password, first_name=full_name)
        user.save()

        # Create StudentProfile
        StudentProfile.objects.create(
            user=user,
            college_name=college_name,
            project_code=project_code
        )

        messages.success(request, 'Account created successfully. Please login.')
        return redirect('login')

    return render(request, 'register.html')

def login_page(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None and not user.is_superuser:
            login(request, user)
            return redirect('dashboard')
        else:
            return render(request, 'login.html', {'error': 'Incorrect credentials.'})

    return render(request, 'login.html')

@login_required(login_url='login')
def dashboard(request):
    
    return render(request, 'dashboard.html')

def logout_user(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def student_leaderboard(request):
    user_profile = StudentProfile.objects.get(user=request.user)
    project_code = user_profile.project_code

    selected_code = request.GET.get('project_code', '')
    sort_by = request.GET.get('sort_by', 'leetcode')
    sort_order = request.GET.get('sort_order', 'desc')

    students = StudentProfile.objects.all()
    if selected_code and selected_code != 'all':
        students = students.filter(project_code=selected_code)

    student_data = []
    for student in students:
        participations = ContestParticipation.objects.filter(student=student)
        gryphon_score = sum(p.total_score for p in participations)
        student_data.append({
            'full_name': student.user.get_full_name() or student.user.username,
            'username': student.user.username,
            'project_code': student.project_code,
            'gryphon_score': round(gryphon_score),
            'leetcode_score': student.leetcode_score,
        })

    reverse = (sort_order == 'desc')
    if sort_by == 'gryphon':
        student_data.sort(key=lambda x: x['gryphon_score'], reverse=reverse)
    else:
        student_data.sort(key=lambda x: x['leetcode_score'], reverse=reverse)

    return render(request, 'student_leaderboard.html', {
        'students': student_data,
        'selected_code': selected_code,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'student_code': project_code
    })

@user_passes_test(is_admin, login_url='admin_login')
def admin_leaderboard(request):
    selected_code = request.GET.get('project_code', '')
    sort_by = request.GET.get('sort_by', 'leetcode')  # default
    sort_order = request.GET.get('sort_order', 'desc')  # 'asc' or 'desc'

    students = StudentProfile.objects.all()
    if selected_code:
        students = students.filter(project_code=selected_code)

    all_codes = StudentProfile.objects.values_list('project_code', flat=True).distinct()
    student_data = []

    for student in students:
        participations = ContestParticipation.objects.filter(student=student)
        gryphon_score = sum(p.total_score for p in participations)
        student_data.append({
            'full_name': student.user.get_full_name() or student.user.username,
            'username': student.user.username,
            'project_code': student.project_code,
            'gryphon_score': round(gryphon_score),
            'leetcode_score': student.leetcode_score,
        })

    reverse = (sort_order == 'desc')
    if sort_by == 'gryphon':
        student_data.sort(key=lambda x: x['gryphon_score'], reverse=reverse)
    else:
        student_data.sort(key=lambda x: x['leetcode_score'], reverse=reverse)

    return render(request, 'admin_leaderboard.html', {
        'students': student_data,
        'all_codes': all_codes,
        'selected_code': selected_code,
        'sort_by': sort_by,
        'sort_order': sort_order
    })

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse

@user_passes_test(is_admin, login_url='admin_login')
def download_leaderboard_excel(request):
    selected_code = request.GET.get('project_code', '')
    sort_by = request.GET.get('sort_by', 'leetcode')
    sort_order = request.GET.get('sort_order', 'desc')

    students = StudentProfile.objects.all()
    if selected_code:
        students = students.filter(project_code=selected_code)

    student_data = []
    for student in students:
        participations = ContestParticipation.objects.filter(student=student)
        gryphon_score = sum(p.total_score for p in participations)
        student_data.append({
            'Name': student.user.get_full_name() or student.user.username,
            'Username': student.user.username,
            'College Name': student.college_name,
            'Gryphon Score': round(gryphon_score),
            'LeetCode Score': student.leetcode_score,
        })

    sort_column = "Gryphon Score" if sort_by == "gryphon" else "LeetCode Score"
    reverse = (sort_order == "desc")
    student_data.sort(key=lambda x: x[sort_column], reverse=reverse)

    # Generate workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Leaderboard"

    headers = list(student_data[0].keys())
    ws.append(headers)

    # Add rows
    for student in student_data:
        ws.append([student[key] for key in headers])

    # Format header row
    for col_num, column_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(column_title) + 6, 18)  # adjust column width
        ws[f"{col_letter}1"].font = Font(bold=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=leaderboard.xlsx'
    wb.save(response)
    return response

@user_passes_test(is_admin, login_url='admin_login')
def admin_dashboard(request):
    print("Logged in user:", request.user)
    print("First name:", request.user.first_name)
    return render(request, 'admin_dashboard.html')

def admin_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_superuser:
            login(request, user)
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid admin credentials.')
    return render(request, 'admin_login.html')

def admin_logout(request):
    logout(request)
    return redirect('admin_login')
